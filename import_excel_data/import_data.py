import mysql.connector
import openpyxl

conn = mysql.connector.connect(host="localhost", user="root", passwd="Vishal#1358", database="Thor")
cur = conn.cursor()

loc = "data.xlsx"

a = openpyxl.load_workbook(loc)
sheet = a.active
values = []


for i in range(2,sheet.max_row+1):
    res = []
    for j in range(1,sheet.max_column+1):
        res.append(sheet.cell(i,j).value)
    values.append(tuple(res))

que = "insert into students (rollno,sname,percentage,branch)values(%s,%s,%s,%s)"

cur.executemany(que,values)
conn.commit()

conn.close()